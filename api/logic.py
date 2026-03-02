import io
import base64
import json
import pandas as pd
import re
from datetime import datetime
import os
# ---------------------------------------------------------------------------
# Configuration & Master Data  (ported from app.py unchanged)
# ---------------------------------------------------------------------------

STANDARD_COLUMNS = [
    "Period", "Location", "Site Type", "Scope", "Category",
    "Fuel / Electricity Type", "Quantity", "Unit", "EF Original Unit",
    "Unit Adjusted EF", "Energy Usage (GJ)", "Total Emissions (kgCO2e)",
    "Total Emissions (tCO2e)", "Factor Source", "Methodology"
]

EF_DATABASE = {
    "Grid Electricity (kWh)":       {"factor": 0.71,    "ncv": 0.0036,  "unit": "kWh", "is_renewable": False, "source": "CEA India 2023",          "methodology": "Scope 2 - Grid Avg",                   "scope": "Scope 2",    "category": "Grid Electricity"},
    "Renewable Electricity (kWh)":  {"factor": 0.0,     "ncv": 0.0036,  "unit": "kWh", "is_renewable": True,  "source": "Renewable",               "methodology": "Zero Emission",                         "scope": "Scope 2",    "category": "Renewable Energy"},
    "Solar (kWh)":                  {"factor": 0.0,     "ncv": 0.0036,  "unit": "kWh", "is_renewable": True,  "source": "Renewable",               "methodology": "Zero Emission",                         "scope": "Scope 2",    "category": "Renewable Energy"},
    "Wind (kWh)":                   {"factor": 0.0,     "ncv": 0.0036,  "unit": "kWh", "is_renewable": True,  "source": "Renewable",               "methodology": "Zero Emission",                         "scope": "Scope 2",    "category": "Renewable Energy"},
    "Hydel (kWh)":                  {"factor": 0.0,     "ncv": 0.0036,  "unit": "kWh", "is_renewable": True,  "source": "Renewable",               "methodology": "Zero Emission",                         "scope": "Scope 2",    "category": "Renewable Energy"},
    "HSD (KL)":                     {"factor": 2701.3,  "ncv": 36.335,  "unit": "KL",  "is_renewable": False, "source": "IPCC 2006, AR5 GWPs",     "methodology": "Scope 1 - Stationary Combustion",       "scope": "Scope 1",    "category": "Stationary Combustion"},
    "Furnace Oil (KL)":             {"factor": 3011.4,  "ncv": 38.784,  "unit": "KL",  "is_renewable": False, "source": "IPCC 2006, AR5 GWPs",     "methodology": "Scope 1 - Stationary Combustion",       "scope": "Scope 1",    "category": "Stationary Combustion"},
    "LDO (KL)":                     {"factor": 2749.3,  "ncv": 36.98,   "unit": "KL",  "is_renewable": False, "source": "IPCC 2006, AR5 GWPs",     "methodology": "Scope 1 - Stationary Combustion",       "scope": "Scope 1",    "category": "Stationary Combustion"},
    "Natural Gas (SCM)":            {"factor": 2.156,   "ncv": 0.0384,  "unit": "SCM", "is_renewable": False, "source": "IPCC 2006, AR5 GWPs",     "methodology": "Scope 1 - Stationary Combustion",       "scope": "Scope 1",    "category": "Stationary Combustion"},
    "LPG (Kg)":                     {"factor": 2.987,   "ncv": 0.0473,  "unit": "kg",  "is_renewable": False, "source": "IPCC 2006, AR5 GWPs",     "methodology": "Scope 1 - Stationary Combustion",       "scope": "Scope 1",    "category": "Stationary Combustion"},
    "LSHS (KL)":                    {"factor": 3042.7,  "ncv": 39.188,  "unit": "KL",  "is_renewable": False, "source": "IPCC 2006, AR5 GWPs",     "methodology": "Scope 1 - Stationary Combustion",       "scope": "Scope 1",    "category": "Stationary Combustion"},
    "HSD Mobile (KL)":              {"factor": 2734.9,  "ncv": 36.335,  "unit": "KL",  "is_renewable": False, "source": "IPCC 2006, AR5 GWPs",     "methodology": "Scope 1 - Mobile Combustion",           "scope": "Scope 1",    "category": "Mobile Combustion"},
    "Petrol (KL)":                  {"factor": 2341.9,  "ncv": 33.004,  "unit": "KL",  "is_renewable": False, "source": "IPCC 2006, AR5 GWPs",     "methodology": "Scope 1 - Mobile Combustion",           "scope": "Scope 1",    "category": "Mobile Combustion"},
    "Biofuel (KL)":                 {"factor": 1688.5,  "ncv": 23.76,   "unit": "KL",  "is_renewable": False, "source": "IPCC 2006, Biogenic",     "methodology": "Biogenic - Reported Separately",        "scope": "Biogenic",   "category": "Biogenic Emissions"},
    "Biodiesel (KL)":               {"factor": 1688.5,  "ncv": 23.76,   "unit": "KL",  "is_renewable": False, "source": "IPCC 2006, Biogenic",     "methodology": "Biogenic - Reported Separately",        "scope": "Biogenic",   "category": "Biogenic Emissions"},
    "Briquettes (Kg)":              {"factor": 1.592,   "ncv": 0.0156,  "unit": "kg",  "is_renewable": False, "source": "IPCC 2006, Biogenic",     "methodology": "Biogenic - Reported Separately",        "scope": "Biogenic",   "category": "Biogenic Emissions"},
    "R22 (kg)":                     {"factor": 1960.0,  "ncv": 0.0,     "unit": "kg",  "is_renewable": False, "source": "IPCC AR6",                "methodology": "Scope 1 - Fugitive Emissions",          "scope": "Scope 1",    "category": "Refrigerants"},
    "R134a (kg)":                   {"factor": 1530.0,  "ncv": 0.0,     "unit": "kg",  "is_renewable": False, "source": "IPCC AR6",                "methodology": "Scope 1 - Fugitive Emissions",          "scope": "Scope 1",    "category": "Refrigerants"},
    "R-407C (kg)":                  {"factor": 1908.0,  "ncv": 0.0,     "unit": "kg",  "is_renewable": False, "source": "IPCC AR6",                "methodology": "Scope 1 - Fugitive Emissions",          "scope": "Scope 1",    "category": "Refrigerants"},
    "R-404A (kg)":                  {"factor": 4728.0,  "ncv": 0.0,     "unit": "kg",  "is_renewable": False, "source": "IPCC AR6",                "methodology": "Scope 1 - Fugitive Emissions",          "scope": "Scope 1",    "category": "Refrigerants"},
    "R-410A (kg)":                  {"factor": 2255.0,  "ncv": 0.0,     "unit": "kg",  "is_renewable": False, "source": "IPCC AR6",                "methodology": "Scope 1 - Fugitive Emissions",          "scope": "Scope 1",    "category": "Refrigerants"},
    "R-407A (kg)":                  {"factor": 2262.0,  "ncv": 0.0,     "unit": "kg",  "is_renewable": False, "source": "IPCC AR6",                "methodology": "Scope 1 - Fugitive Emissions",          "scope": "Scope 1",    "category": "Refrigerants"},
    "R123 (kg)":                    {"factor": 79.0,    "ncv": 0.0,     "unit": "kg",  "is_renewable": False, "source": "IPCC AR6",                "methodology": "Scope 1 - Fugitive Emissions",          "scope": "Scope 1",    "category": "Refrigerants"},
    "R32 (kg)":                     {"factor": 771.0,   "ncv": 0.0,     "unit": "kg",  "is_renewable": False, "source": "IPCC AR6",                "methodology": "Scope 1 - Fugitive Emissions",          "scope": "Scope 1",    "category": "Refrigerants"},
    "R152a (kg)":                   {"factor": 164.0,   "ncv": 0.0,     "unit": "kg",  "is_renewable": False, "source": "IPCC AR6",                "methodology": "Scope 1 - Fugitive Emissions",          "scope": "Scope 1",    "category": "Refrigerants"},
    "Fire Extinguisher 1kg (nos)":    {"factor": 1.0,   "ncv": 0.0,     "unit": "nos", "is_renewable": False, "source": "CO2 Volume",              "methodology": "Scope 1 - Fugitive Emissions",          "scope": "Scope 1",    "category": "Fire Extinguishers"},
    "Fire Extinguisher 2kg (nos)":    {"factor": 2.0,   "ncv": 0.0,     "unit": "nos", "is_renewable": False, "source": "CO2 Volume",              "methodology": "Scope 1 - Fugitive Emissions",          "scope": "Scope 1",    "category": "Fire Extinguishers"},
    "Fire Extinguisher 2.5kg (nos)":  {"factor": 2.5,   "ncv": 0.0,     "unit": "nos", "is_renewable": False, "source": "CO2 Volume",              "methodology": "Scope 1 - Fugitive Emissions",          "scope": "Scope 1",    "category": "Fire Extinguishers"},
    "Fire Extinguisher 3kg (nos)":    {"factor": 3.0,   "ncv": 0.0,     "unit": "nos", "is_renewable": False, "source": "CO2 Volume",              "methodology": "Scope 1 - Fugitive Emissions",          "scope": "Scope 1",    "category": "Fire Extinguishers"},
    "Fire Extinguisher 3.5kg (nos)":  {"factor": 3.5,   "ncv": 0.0,     "unit": "nos", "is_renewable": False, "source": "CO2 Volume",              "methodology": "Scope 1 - Fugitive Emissions",          "scope": "Scope 1",    "category": "Fire Extinguishers"},
    "Fire Extinguisher 4.5kg (nos)":  {"factor": 4.5,   "ncv": 0.0,     "unit": "nos", "is_renewable": False, "source": "CO2 Volume",              "methodology": "Scope 1 - Fugitive Emissions",          "scope": "Scope 1",    "category": "Fire Extinguishers"},
    "Fire Extinguisher 6kg (nos)":    {"factor": 6.0,   "ncv": 0.0,     "unit": "nos", "is_renewable": False, "source": "CO2 Volume",              "methodology": "Scope 1 - Fugitive Emissions",          "scope": "Scope 1",    "category": "Fire Extinguishers"},
    "Fire Extinguisher 9kg (nos)":    {"factor": 9.0,   "ncv": 0.0,     "unit": "nos", "is_renewable": False, "source": "CO2 Volume",              "methodology": "Scope 1 - Fugitive Emissions",          "scope": "Scope 1",    "category": "Fire Extinguishers"},
    "Fire Extinguisher 22.5kg (nos)": {"factor": 22.5,  "ncv": 0.0,     "unit": "nos", "is_renewable": False, "source": "CO2 Volume",              "methodology": "Scope 1 - Fugitive Emissions",          "scope": "Scope 1",    "category": "Fire Extinguishers"},
    "Coal (Kg)":                      {"factor": 1.46,  "ncv": 0.015,   "unit": "kg",  "is_renewable": False, "source": "IPCC 2006, AR5 GWPs",     "methodology": "Scope 1 - Stationary Combustion",       "scope": "Scope 1",    "category": "Stationary Combustion"},
}

FUEL_MAPPING = {
    "electricity": "Grid Electricity (kWh)", "kwh": "Grid Electricity (kWh)", "power": "Grid Electricity (kWh)", "nre": "Grid Electricity (kWh)",
    "solar": "Solar (kWh)", "wind": "Wind (kWh)", "hydel": "Hydel (kWh)", "hydro": "Hydel (kWh)", "hydroelectric": "Hydel (kWh)",
    "re": "Renewable Electricity (kWh)", "renewable": "Renewable Electricity (kWh)", "ppa": "Renewable Electricity (kWh)", "rooftop": "Renewable Electricity (kWh)", "open access": "Renewable Electricity (kWh)",
    "hsd": "HSD (KL)", "diesel": "HSD (KL)", "natural gas": "Natural Gas (SCM)", "ng": "Natural Gas (SCM)",
    "lpg": "LPG (Kg)", "fo": "Furnace Oil (KL)", "furnace oil": "Furnace Oil (KL)", "furnance oil": "Furnace Oil (KL)",
    "ldo": "LDO (KL)", "lshs": "LSHS (KL)", "petrol": "Petrol (KL)",
    "biofuel": "Biofuel (KL)", "biodiesel": "Biodiesel (KL)", "briquettes": "Briquettes (Kg)",
    "r152a": "R152a (kg)", "r-152a": "R152a (kg)",
    "r22": "R22 (kg)", "r134a": "R134a (kg)", "r-134a": "R134a (kg)", "r134": "R134a (kg)",
    "r407c": "R-407C (kg)", "r-407c": "R-407C (kg)", "r407": "R-407C (kg)",
    "r404a": "R-404A (kg)", "r-404a": "R-404A (kg)", "r404": "R-404A (kg)",
    "r410a": "R-410A (kg)", "r-410a": "R-410A (kg)", "r410": "R-410A (kg)",
    "r407a": "R-407A (kg)", "r-407a": "R-407A (kg)",
    "r123": "R123 (kg)", "r-123": "R123 (kg)",
    "r32": "R32 (kg)", "r-32": "R32 (kg)",
    "png": "Natural Gas (SCM)",
    "piped natural gas": "Natural Gas (SCM)",
    "cng": "Natural Gas (SCM)",
    "compressed natural gas": "Natural Gas (SCM)",
    "coal": "Coal (Kg)",
    "bituminous coal": "Coal (Kg)",
    "sub-bituminous coal": "Coal (Kg)",
    "hsd (mobile)": "HSD Mobile (KL)",
    "hsd mobile": "HSD Mobile (KL)",
    "fire extinguisher refilled: 1 kg":    "Fire Extinguisher 1kg (nos)",
    "fire extinguisher refilled: 2 kg":    "Fire Extinguisher 2kg (nos)",
    "fire extinguisher refilled: 2.5 kg":  "Fire Extinguisher 2.5kg (nos)",
    "fire extinguisher refilled: 3 kg":    "Fire Extinguisher 3kg (nos)",
    "fire extinguisher refilled: 3.5 kg":  "Fire Extinguisher 3.5kg (nos)",
    "fire extinguisher refilled: 4.5 kg":  "Fire Extinguisher 4.5kg (nos)",
    "fire extinguisher refilled: 6 kg":    "Fire Extinguisher 6kg (nos)",
    "fire extinguisher refilled: 9 kg":    "Fire Extinguisher 9kg (nos)",
    "fire extinguisher refilled: 22.5 kg": "Fire Extinguisher 22.5kg (nos)",
}

INDIAN_CITIES = [
    'nashik', 'baddi', 'goa', 'indore', 'sikkim', 'mumbai', 'delhi', 'bangalore',
    'chennai', 'hyderabad', 'pune', 'gurgaon', 'noida', 'ahmedabad', 'kolkata',
    'nalagarh', 'dindori', 'aurangabad', 'mahape', 'sinnar', 'taloja', 'sanpada', 'mohol'
]
EXCLUDE_KW = [
    "total", "sum", "emission factor", "emissions tco2", "tco2e", "tco2", "ncv",
    "intensity", "source", "methodology", "scope", "category", "density", "gwp",
    "parameter", "notes", "co2e emission"
]

FUEL_KEYS_SORTED = sorted(FUEL_MAPPING.keys(), key=len, reverse=True)

# ---------------------------------------------------------------------------
# Intelligence Engine  (ported unchanged)
# ---------------------------------------------------------------------------

def safe_float(v):
    if pd.isna(v) or v is None:
        return 0.0
    clean_v = re.sub(r'[^\d\.-]', '', str(v).strip())
    try:
        return float(clean_v) if clean_v else 0.0
    except Exception:
        return 0.0


def map_fuel_name(text, default="Unknown"):
    t = str(text).strip().lower()
    if not t:
        return default
    if re.search(r"(?<![a-z0-9])nre(?![a-z0-9])", t):
        return "Grid Electricity (kWh)"
    if re.search(r"(?<![a-z0-9])(hydel|hydro|hydroelectric)(?![a-z0-9])", t):
        return "Hydel (kWh)"
    if re.search(r"(?<![a-z0-9])solar(?![a-z0-9])", t):
        return "Solar (kWh)"
    if re.search(r"(?<![a-z0-9])wind(?![a-z0-9])", t):
        return "Wind (kWh)"
    if re.search(r"(?<![a-z0-9])(re|renewable|open access|rooftop|ppa)(?![a-z0-9])", t):
        return "Renewable Electricity (kWh)"
    for kw in FUEL_KEYS_SORTED:
        pattern = rf"(?<![a-z0-9]){re.escape(kw)}(?![a-z0-9])"
        if re.search(pattern, t):
            return FUEL_MAPPING[kw]
    for kw in FUEL_KEYS_SORTED:
        if kw in t:
            return FUEL_MAPPING[kw]
    return default


def is_activity_location(loc):
    l = str(loc).strip().lower()
    if l in ["", "nan", "none", "null", "n/a", "-"]:
        return False
    bad_tokens = ["parameter", "source", "density", "ncv", "gwp", "co2", "ch4", "n2o",
                  "emission factor", "notes"]
    # "site" excluded from bad_tokens: substring match wrongly rejects "Site-Alpha" etc.
    if l in ["site", "site name", "location", "plant name"]:
        return False
    if any(tok in l for tok in bad_tokens):
        return False
    return True


def is_reference_text(text):
    t = str(text).lower()
    ref_tokens = ["emission factors", "parameter", "density", "ncv", "gwp",
                  "co2 ef", "ch4 ef", "n2o ef", "source", "defra", "ipcc", "tco2e", "co2e emissions"]
    return any(tok in t for tok in ref_tokens)


def is_valid_reporting_period(period_text):
    p = str(period_text).strip().upper()
    if not p:
        return False
    if "FY " in p and re.search(r"20\d{2}\s*[-/]\s*\d{2,4}", p):
        return True
    if re.search(r"\b20\d{2}\s*[-/]\s*\d{2,4}\b", p):
        return True
    return False


def extract_custom_ef_from_header(text):
    t = str(text).lower()
    if "custom" not in t:
        return None
    patterns = [
        r"(?:ef|emission\s*factor)\s*[:=]?\s*(-?\d+(?:\.\d+)?)",
        r"(-?\d+(?:\.\d+)?)\s*(?:kg\s*co2e\s*/\s*kwh|kgco2e\s*/\s*kwh|co2e\s*/\s*kwh)",
        r"(-?\d+(?:\.\d+)?)\s*(?:t\s*co2e\s*/\s*mwh|tco2e\s*/\s*mwh)",
        r"(-?\d+(?:\.\d+)?)\s*(?:g\s*co2e\s*/\s*kwh|gco2e\s*/\s*kwh)"
    ]
    for pat in patterns:
        m = re.search(pat, t)
        if m:
            val = safe_float(m.group(1))
            if "gco2e" in pat or "g\\s*co2e" in pat:
                return val / 1000.0
            return val
    return None


def resolve_fuel_profile(raw_text, mapped_fuel):
    t = str(raw_text).lower()
    f = mapped_fuel
    is_re_context = any(k in t for k in ["re", "renewable", "solar", "wind", "hydel", "hydro",
                                          "green power", "open access", "rooftop", "ppa"]) or \
                    ("custom" in t and "kwh" in t)
    custom_ef = extract_custom_ef_from_header(raw_text)
    if custom_ef is not None and is_re_context:
        return (
            "Custom Renewable Electricity (kWh)",
            {
                "factor": custom_ef, "ncv": 0.0036, "unit": "kWh", "is_renewable": True,
                "source": "Custom EF from header", "methodology": "Scope 2 - Custom Renewable",
                "scope": "Scope 2", "category": "Renewable Energy"
            }
        )
    return f, None


def find_unit(text):
    text_lower = str(text).lower()
    mapping = {
        r'\bkg\b|\bkilogram\b': 'kg',
        r'\bkl\b|\bkilolitre\b|\bkiloliter\b': 'KL',
        r'\bltr\b|\blitre\b|\bliter\b|\bliters\b|\blitres\b': 'Litre',
        r'\bsch?m\b|\bscm\b|\bsm3\b|\bsm³\b': 'SCM',
        r'\bkwh\b|\bunit\b': 'kWh',
        r'\bmt\b|\bmetric tonn?e\b': 'MT'
    }
    for pattern, unit in mapping.items():
        if re.search(pattern, text_lower):
            return unit
    return "N/A"


def extract_period_metadata(text):
    t = str(text).lower()
    fy_match = re.search(r"(?:fy)?\s*(20\d{2}[\-\/]\d{2,4}|\d{2}[\-\/]\d{2})", t)
    if fy_match:
        return f"FY {fy_match.group(1)}".upper()
    month_match = re.search(r"([a-z]{3,9})[\s\-\/]*(\d{2,4})", t)
    if month_match:
        m_str, y_str = month_match.groups()
        try:
            m_name = datetime.strptime(m_str[:3], "%b").strftime("%B")
            y_name = f"20{y_str}" if len(y_str) == 2 else y_str
            return f"{m_name} {y_name}"
        except Exception:
            pass
    year_match = re.search(r"\b(20\d{2})\b", t)
    if year_match:
        return year_match.group(1)
    return ""


def validate_location(loc):
    if not loc or pd.isna(loc):
        return "Unknown Site"
    l_str = str(loc).lower().strip()
    if l_str in ['nan', 'none', 'null', 'unknown', 'n/a', '', '-']:
        return "Unknown Site"
    if any(x == l_str for x in EXCLUDE_KW):
        return "Unknown Site"
    return str(loc).strip()


def detect_site_type(name):
    name_low = str(name).lower()
    if any(x in name_low for x in ["plant", "factory", "mfg", "unit", "work", "prod"]):
        return "Manufacturing Site"
    if any(x in name_low for x in ["warehouse", "depot", "wh", "store"]):
        return "Warehouse"
    if any(x in name_low for x in ["office", "hq", "corp", "tower", "suite"]):
        return "Office"
    if any(x in name_low for x in ["village", "rural", "gram"]):
        return "Village/Rural Site"
    return "Commercial Site"


def detect_and_melt_matrix(df):
    loc_col = None
    for col in df.columns:
        if any(city in df[col].astype(str).str.lower().tolist() for city in INDIAN_CITIES):
            loc_col = col
            break
    if not loc_col:
        return df, False, None
    fuel_headers = [
        c for c in df.columns
        if c != loc_col
        and map_fuel_name(c, default=None) is not None
        and not is_reference_text(c)
    ]
    if any(extract_period_metadata(c) for c in fuel_headers):
        fuel_headers = [c for c in fuel_headers if extract_period_metadata(c)]
    if len(fuel_headers) < 2:
        return df, False, None
    value_vars = [c for c in fuel_headers if not any(kw in str(c).lower() for kw in EXCLUDE_KW)]
    id_vars = [loc_col]
    for c in df.columns:
        if c != loc_col and any(kw in str(c).lower() for kw in ["month", "year", "fy", "period"]) and c not in value_vars:
            id_vars.append(c)
    melted = pd.melt(df, id_vars=id_vars, value_vars=value_vars,
                     var_name='Fuel / Electricity Type', value_name='Quantity')
    melted = melted.dropna(subset=['Quantity'])
    melted['Quantity'] = melted['Quantity'].apply(safe_float)
    return melted[melted['Quantity'] > 0], True, loc_col


def split_sheet_into_tables(df):
    mask = df.isna().all(axis=1)
    is_gap = mask & mask.shift(1, fill_value=False)
    groups = is_gap.cumsum()
    sub_dfs = [group_df for _, group_df in df[~mask].groupby(groups[~mask]) if len(group_df) > 1]
    return sub_dfs if sub_dfs else ([df[~mask]] if len(df[~mask]) > 1 else [])


def try_parse_tidy(raw_df, default_period="", sheet_context=""):
    """Detect and process tidy/long-row format where each row = one observation.

    Handles sheets like:
      FY       | Month    | Site       | Fuel | Unit | Quantity
      FY2024-25| Apr-2024 | Site-Alpha | PNG  | Sm3  | 15918.92
    or
      FY       | Month    | Site       | Consumption_kWh
      FY2024-25| Apr-2024 | Site-Alpha | 171589

    Returns a list of processed rows, or [] if this format is not detected.
    """
    df = raw_df.copy()

    # 1. Find the header row (scan first 5 rows for tidy column keywords)
    _SITE_KW   = ["site", "location", "plant", "branch", "area"]
    _FUEL_KW   = ["fuel", "energy type", "fuel type", "parameter"]
    _QTY_KW    = ["quantity", "qty", "amount", "consumption", "volume"]
    _UNIT_KW   = ["unit", "uom", "units"]
    _PERIOD_KW = ["month", "date", "period"]
    _FY_KW     = ["fy", "financial year", "fiscal year"]

    header_row_idx = None
    for idx in range(min(5, len(df))):
        vals = [str(v).strip().lower() if pd.notna(v) else "" for v in df.iloc[idx].tolist()]
        has_site  = any(kw in v for v in vals for kw in _SITE_KW)
        has_fuel  = any(kw == v for v in vals for kw in _FUEL_KW) or \
                    any(v == "fuel" for v in vals)
        has_qty   = any(kw in v for v in vals for kw in _QTY_KW)
        if has_site and (has_fuel or has_qty):
            header_row_idx = idx
            break

    if header_row_idx is None:
        return []

    # 2. Apply header row as column names
    raw_cols = [str(v).strip() if pd.notna(v) else f"_col_{i}"
                for i, v in enumerate(df.iloc[header_row_idx].tolist())]
    data_df = df.iloc[header_row_idx + 1:].reset_index(drop=True)
    data_df.columns = raw_cols
    data_df = data_df.dropna(how='all').reset_index(drop=True)
    if data_df.empty:
        return []

    col_lower = {c: c.lower() for c in raw_cols}

    # 3. Identify columns by role
    fuel_col = next((c for c in raw_cols
                     if col_lower[c] in ["fuel", "fuel type", "energy type", "type", "parameter"]), None)
    site_col = next((c for c in raw_cols
                     if any(kw in col_lower[c] for kw in _SITE_KW)), None)
    unit_col = next((c for c in raw_cols
                     if any(kw in col_lower[c] for kw in _UNIT_KW)
                     and "quantity" not in col_lower[c]), None)
    fy_col   = next((c for c in raw_cols
                     if col_lower[c] in ["fy", "financial year", "fiscal year"]), None)
    month_col = next((c for c in raw_cols
                      if any(kw in col_lower[c] for kw in _PERIOD_KW)
                      and c != fy_col), None)

    # Quantity col: named quantity/consumption, or a single numeric non-id column
    qty_col = None
    skip_cols = {c for c in [fuel_col, site_col, unit_col, fy_col, month_col] if c}
    for c in raw_cols:
        if c in skip_cols:
            continue
        cl = col_lower[c]
        if any(kw in cl for kw in _QTY_KW):
            if pd.to_numeric(data_df[c], errors='coerce').notna().any():
                qty_col = c
                break
    if qty_col is None:
        for c in raw_cols:
            if c in skip_cols:
                continue
            converted = pd.to_numeric(data_df[c], errors='coerce')
            if converted.notna().sum() > len(data_df) * 0.4:
                qty_col = c
                break

    if site_col is None or qty_col is None:
        return []

    # Electricity-only: no fuel col but qty col name contains "kwh"
    is_elec_only = (fuel_col is None and "kwh" in col_lower.get(qty_col, ""))

    if fuel_col is None and not is_elec_only:
        return []

    # 4. Process rows
    results = []
    for _, row in data_df.iterrows():
        # Period: prefer FY column (gives "FY 2024-25") over Month ("April 2024")
        period = default_period
        if fy_col and pd.notna(row.get(fy_col)):
            extracted = extract_period_metadata(str(row[fy_col]))
            if extracted:
                period = extracted
        if not is_valid_reporting_period(period) and month_col and pd.notna(row.get(month_col)):
            extracted = extract_period_metadata(str(row[month_col]))
            if extracted and is_valid_reporting_period(extracted):
                period = extracted
        if not is_valid_reporting_period(period):
            continue

        # Location
        loc_val = row.get(site_col)
        loc = str(loc_val).strip() if pd.notna(loc_val) else ""
        if not loc or loc.lower() in ["", "nan", "none", "null", "n/a", "-"]:
            continue

        # Quantity
        qty = safe_float(row.get(qty_col, 0))
        if qty <= 0:
            continue

        # Fuel type & unit
        if is_elec_only:
            f_type = "Grid Electricity (kWh)"
            unit = "kWh"
        else:
            raw_fuel = str(row.get(fuel_col, "")).strip() if pd.notna(row.get(fuel_col)) else ""
            if not raw_fuel or raw_fuel.lower() in ["", "nan", "none"]:
                continue
            f_type = map_fuel_name(raw_fuel, default="Unknown")
            raw_unit = str(row.get(unit_col, "")).strip() if (unit_col and pd.notna(row.get(unit_col))) else ""
            unit = find_unit(raw_unit) if raw_unit else find_unit(raw_fuel)

        results.append(process_standard_row(f_type, qty, period, loc, unit))

    return results


def process_electricity_sheet(raw_df, default_period=""):
    rows = []
    df = raw_df.copy()
    location_fallback_col = 0

    def _is_valid_site_label(site_text):
        s = str(site_text).strip().lower()
        if s in ["", "nan", "none", "null", "n/a", "-"]:
            return False
        bad_tokens = ["total", "energy gj", "energy renewable", "energy non renewable",
                      "emissions", "re %", "scope 2 -", "tco2", "intensity"]
        if any(tok in s for tok in bad_tokens):
            return False
        return True

    for r in range(len(df)):
        row_vals = [str(v).strip().lower() if pd.notna(v) else "" for v in df.iloc[r].tolist()]
        has_nre = "nre" in row_vals
        has_re = "re" in row_vals
        if not (has_nre and has_re):
            continue
        year_row_idx = max(r - 1, 0)
        year_vals = [str(v) if pd.notna(v) else "" for v in df.iloc[year_row_idx].tolist()]
        for c in range(len(row_vals)):
            token = row_vals[c]
            if token not in ["kwh", "nre"]:
                continue
            if token == "kwh":
                nre_col = c + 1 if c + 1 < len(row_vals) and row_vals[c + 1] == "nre" else None
                re_col = c + 2 if c + 2 < len(row_vals) and row_vals[c + 2] == "re" else None
                site_col = c - 1
            else:
                nre_col = c
                re_col = c + 1 if c + 1 < len(row_vals) and row_vals[c + 1] == "re" else None
                site_col = c - 2
            if nre_col is None:
                continue
            period = ""
            for cc in [nre_col, nre_col - 1, nre_col + 1, nre_col - 2, nre_col + 2]:
                if 0 <= cc < len(year_vals):
                    period = extract_period_metadata(year_vals[cc]) or period
            period = period or default_period
            for rr in range(r + 1, len(df)):
                first_txt = str(df.iloc[rr, 0]).strip().lower() if pd.notna(df.iloc[rr, 0]) else ""
                if first_txt in ["", "nan"]:
                    continue
                if any(k in first_txt for k in ["total", "energy gj", "emissions", "re %", "scope 2 -"]):
                    break
                site_val = (df.iloc[rr, site_col]
                            if 0 <= site_col < df.shape[1] and pd.notna(df.iloc[rr, site_col])
                            else df.iloc[rr, location_fallback_col])
                site = str(site_val).strip()
                if not is_activity_location(site) or not _is_valid_site_label(site):
                    continue
                nre_qty = safe_float(df.iloc[rr, nre_col]) if 0 <= nre_col < df.shape[1] else 0.0
                re_qty = safe_float(df.iloc[rr, re_col]) if re_col is not None and 0 <= re_col < df.shape[1] else 0.0
                if nre_qty > 0 and is_valid_reporting_period(period):
                    rows.append(process_standard_row("Grid Electricity (kWh)", nre_qty, period, site, "kWh"))
                if re_qty > 0 and is_valid_reporting_period(period):
                    rows.append(process_standard_row("Renewable Electricity (kWh)", re_qty, period, site, "kWh"))
    if not rows:
        return rows
    seen = set()
    dedup = []
    for row in rows:
        key = (row["Period"], row["Location"], row["Fuel / Electricity Type"], round(row["Quantity"], 6))
        if key in seen:
            continue
        seen.add(key)
        dedup.append(row)
    return dedup


def process_standard_row(f_type, qty, period="", loc="Unknown", u_in="N/A", ef_override=None):
    q_val = safe_float(qty)
    ef = EF_DATABASE.get(f_type, {
        "factor": 0, "ncv": 0, "unit": "N/A", "scope": "N/A",
        "category": "N/A", "is_renewable": False
    }).copy()
    if ef_override:
        ef.update(ef_override)
    adj = 0.001 if (u_in.lower() in ["litre", "ltr", "liter"] and ef['unit'] == "KL") else 1.0
    adj_ef = ef['factor'] * adj
    kg_co2 = q_val * adj_ef
    return {
        "Period": period,
        "Location": validate_location(loc),
        "Site Type": detect_site_type(loc),
        "Scope": ef.get('scope', 'N/A'),
        "Category": ef.get('category', 'N/A'),
        "Fuel / Electricity Type": f_type,
        "Quantity": q_val,
        "Unit": u_in if u_in != "N/A" else ef.get('unit', 'N/A'),
        "EF Original Unit": ef.get('unit', 'N/A'),
        "Unit Adjusted EF": adj_ef,
        "Energy Usage (GJ)": q_val * adj * ef['ncv'],
        "Total Emissions (kgCO2e)": kg_co2,
        "Total Emissions (tCO2e)": kg_co2 / 1000,
        "Factor Source": ef.get('source', 'N/A'),
        "Methodology": ef.get('methodology', 'N/A'),
    }


def energy_gj_from_row(row):
    qty = safe_float(row.get("Quantity", 0))
    f_type = str(row.get("Fuel / Electricity Type", ""))
    unit = str(row.get("Unit", "")).strip().lower()
    ef = EF_DATABASE.get(f_type, {"unit": "N/A", "ncv": 0.0})
    ef_unit = str(ef.get("unit", "")).strip().lower()
    ncv = safe_float(ef.get("ncv", 0.0))
    if unit in ["gj"]:
        return qty
    if unit in ["mwh"]:
        return qty * 3.6
    if unit in ["kwh", "unit"]:
        return qty * 0.0036
    if unit in ["l", "ltr", "litre", "liter"] and ef_unit == "kl":
        return qty * 0.001 * ncv
    if unit == "kl" and ef_unit == "kl":
        return qty * ncv
    if unit in ["kg", "kilogram"] and ef_unit == "kg":
        return qty * ncv
    if unit == "scm" and ef_unit == "scm":
        return qty * ncv
    return safe_float(row.get("Energy Usage (GJ)", 0.0))


def classify_headers(df):
    cl = {"location": [], "period": None, "fuels": {}, "qty": [], "other": []}
    sample_blob = " ".join(str(v) for v in df.astype(str).head(10).values.flatten().tolist()).lower()
    guessed_fuel = map_fuel_name(sample_blob, default=None)

    for col in df.columns:
        c_low = str(col).lower()
        if any(kw in c_low for kw in EXCLUDE_KW):
            continue
        if is_reference_text(col):
            continue
        is_num = pd.api.types.is_numeric_dtype(df[col])
        p_ext = extract_period_metadata(col)
        fuel = map_fuel_name(c_low, default=None)
        if p_ext and is_num and fuel:
            cl["fuels"][col] = fuel
            continue
        if p_ext:
            if is_num:
                cl["fuels"][col] = "Date-Mapped Quantity"
            else:
                cl["period"] = col
            continue
        if any(kw in c_low for kw in ["month", "year", "fy", "period"]) and "qty" not in c_low:
            cl["period"] = col
            continue
        if fuel:
            cl["fuels"][col] = fuel
            continue
        if any(x in c_low for x in ["site", "location", "plant", "branch", "area"]):
            cl["location"].append(col)
            continue
        if is_num:
            cl["qty"].append({"name": col, "fuel": guessed_fuel})

    if any(extract_period_metadata(c) for c in cl["fuels"].keys()):
        cl["fuels"] = {c: f for c, f in cl["fuels"].items() if extract_period_metadata(c)}
    return cl, df


def process_table_block(df, parent_period, sheet_context=""):
    if "consolidated" in sheet_context:
        return []
    is_mobile = "mobile" in sheet_context
    stop_re = re.compile(r"co.?2e emissions|emission factors|total scope 1 emissions|notes:", re.IGNORECASE)
    cutoff = None
    for idx in range(len(df)):
        preview = " ".join(str(x) for x in df.iloc[idx, :6].tolist() if pd.notna(x)).lower()
        if idx > 5 and stop_re.search(preview):
            cutoff = idx
            break
    if cutoff is not None:
        df = df.iloc[:cutoff].reset_index(drop=True)

    df = df.dropna(how='all', axis=1)
    df.columns = [str(c) for c in df.columns]
    if df.empty:
        return []

    top_blob = " ".join(
        str(v) for v in df.head(3).astype(str).values.flatten().tolist()
        if str(v).strip() and str(v).lower() != "nan"
    ).lower()
    if re.search(r"co.?2e emissions|emission factors|scope 1 consolidated ghg emissions", top_blob):
        return []

    header_keywords = [kw for kw in FUEL_MAPPING.keys() if len(kw) >= 3] + [
        "date", "period", "loc", "site", "year", "month", "qty", "unit",
        "location", "quantity", "consumption"
    ]
    header_patterns = [re.compile(rf'\b{re.escape(kw)}\b', re.IGNORECASE) for kw in header_keywords]
    best_row, max_score = 0, -1
    for idx in range(min(10, len(df))):
        row = df.iloc[idx].astype(str).str.lower()
        score = sum(1 for cell in row if any(p.search(str(cell)) for p in header_patterns))
        if score > max_score:
            max_score, best_row = score, idx

    if max_score > 0:
        new_cols = df.iloc[best_row].tolist()
        period_row = [None] * len(new_cols)
        for prev_idx in range(best_row):
            prev_vals = df.iloc[prev_idx].tolist()
            last_period = None
            for ci, cell in enumerate(prev_vals):
                p = extract_period_metadata(str(cell)) if pd.notna(cell) else None
                if p:
                    last_period = p
                if last_period:
                    period_row[ci] = last_period
        if not parent_period:
            parent_period = next((p for p in period_row if p), "")

        clean_cols = []
        has_multi_periods = len(set(p for p in period_row if p)) > 1
        for i, c in enumerate(new_cols):
            col_name = str(c).strip() if pd.notna(c) else f"_col_{i}"
            if has_multi_periods and period_row[i]:
                col_name = f"{period_row[i]} | {col_name}"
            clean_cols.append(col_name)

        seen = {}
        for i, c in enumerate(clean_cols):
            if c in seen:
                seen[c] += 1
                clean_cols[i] = f"{c}_{seen[c]}"
            else:
                seen[c] = 0
        df.columns = clean_cols
        df = df.iloc[best_row + 1:].reset_index(drop=True)

    def _is_total_row(r):
        try:
            return r.astype(str).str.contains(
                r'total|sum|grand|subtotal|emission.?factor|emissions?\s*tco2|tco2e|density|gwp|parameter|co2e.?emission|notes:|ipcc|ghg protocol',
                case=False, na=False
            ).any()
        except Exception:
            return False

    df = df[~df.apply(_is_total_row, axis=1)]

    m_df, is_matrix, loc_col = detect_and_melt_matrix(df)

    def _resolve_fuel(f_type):
        if is_mobile and f_type in ["HSD (KL)", "HSD (Mobile)"]:
            return "HSD Mobile (KL)"
        return f_type

    results = []
    if is_matrix:
        for _, r in m_df.iterrows():
            loc = str(r[loc_col])
            if not is_activity_location(loc):
                continue
            f_t_raw = str(r['Fuel / Electricity Type'])
            f_type = map_fuel_name(f_t_raw, default="Unknown")
            f_type = _resolve_fuel(f_type)
            f_type, ef_override = resolve_fuel_profile(f_t_raw, f_type)
            val = safe_float(r['Quantity'])
            row_p = extract_period_metadata(f_t_raw) or parent_period
            period_cols = [c for c in r.index
                           if any(kw in str(c).lower() for kw in ["month", "year", "fy", "period"])
                           and c != 'Fuel / Electricity Type']
            if period_cols and pd.notna(r[period_cols[0]]):
                row_p = extract_period_metadata(r[period_cols[0]]) or row_p
            if val > 0 and is_valid_reporting_period(row_p):
                results.append(process_standard_row(f_type, val, row_p, loc, find_unit(f_t_raw), ef_override))
    else:
        cl, df = classify_headers(df)
        for _, r in df.iterrows():
            row_p = parent_period
            if cl["period"] and pd.notna(r[cl["period"]]):
                row_p = extract_period_metadata(r[cl["period"]]) or row_p
            loc = str(r[cl['location'][0]]) if cl['location'] else "Unknown Site"
            if not is_activity_location(loc):
                continue
            for c, f_t in cl["fuels"].items():
                val = safe_float(r[c])
                if val > 0:
                    cur_p = extract_period_metadata(c) or row_p if f_t == "Date-Mapped Quantity" else row_p
                    if is_valid_reporting_period(cur_p):
                        resolved_fuel = _resolve_fuel(f_t)
                        resolved_fuel, ef_override = resolve_fuel_profile(c, resolved_fuel)
                        results.append(process_standard_row(resolved_fuel, val, cur_p, loc, find_unit(c), ef_override))
            if not cl["fuels"]:
                for q in cl["qty"]:
                    val = safe_float(r[q['name']])
                    if val > 0:
                        if is_valid_reporting_period(row_p):
                            resolved_fuel = _resolve_fuel(q['fuel'] or "Unknown")
                            resolved_fuel, ef_override = resolve_fuel_profile(q['name'], resolved_fuel)
                            results.append(process_standard_row(resolved_fuel, val, row_p, loc, find_unit(q['name']), ef_override))
    return results


def get_fy_start(period_text):
    m = re.search(r"FY\s*(20\d{2})\s*[-/]\s*(\d{2,4})", str(period_text), re.IGNORECASE)
    if not m:
        return None
    return int(m.group(1))


def format_fy(start_year):
    return f"FY {start_year}-{str(start_year + 1)[-2:]}"


def build_yearly_summary_with_proxy(rdf, proxy_years_if_single=2, decline=0.05):
    df = rdf.copy()
    df["FY Start"] = df["Period"].apply(get_fy_start)
    df = df[df["FY Start"].notna()].copy()
    if df.empty:
        return pd.DataFrame()

    grouped = df.groupby("FY Start", as_index=False).agg({
        "Total Emissions (tCO2e)": "sum",
        "Energy Usage (GJ)": "sum"
    })
    s1 = df[df["Scope"] == "Scope 1"].groupby("FY Start")["Total Emissions (tCO2e)"].sum()
    s2 = df[df["Scope"] == "Scope 2"].groupby("FY Start")["Total Emissions (tCO2e)"].sum()
    grouped["Scope 1"] = grouped["FY Start"].map(s1).fillna(0.0)
    grouped["Scope 2"] = grouped["FY Start"].map(s2).fillna(0.0)
    grouped["Data Type"] = "Actual"

    rows = {int(r["FY Start"]): r.to_dict() for _, r in grouped.iterrows()}
    actual_years = sorted(rows.keys())
    min_year, max_year = min(actual_years), max(actual_years)

    for y in range(max_year - 1, min_year - 1, -1):
        if y not in rows and (y + 1) in rows:
            src = rows[y + 1]
            rows[y] = {
                "FY Start": y,
                "Total Emissions (tCO2e)": src["Total Emissions (tCO2e)"] * (1 - decline),
                "Energy Usage (GJ)": src["Energy Usage (GJ)"] * (1 - decline),
                "Scope 1": src["Scope 1"] * (1 - decline),
                "Scope 2": src["Scope 2"] * (1 - decline),
                "Data Type": "Proxy"
            }

    if len(actual_years) == 1:
        latest = actual_years[0]
        for i in range(1, proxy_years_if_single + 1):
            y = latest - i
            src = rows[y + 1]
            rows[y] = {
                "FY Start": y,
                "Total Emissions (tCO2e)": src["Total Emissions (tCO2e)"] * (1 - decline),
                "Energy Usage (GJ)": src["Energy Usage (GJ)"] * (1 - decline),
                "Scope 1": src["Scope 1"] * (1 - decline),
                "Scope 2": src["Scope 2"] * (1 - decline),
                "Data Type": "Proxy"
            }

    out = pd.DataFrame(rows.values()).sort_values("FY Start").reset_index(drop=True)
    out["Period"] = out["FY Start"].apply(format_fy)
    return out


# ---------------------------------------------------------------------------
# Flask API routes
# ---------------------------------------------------------------------------

RENEWABLE_ELEC_TYPES = {
    "Renewable Electricity (kWh)", "Solar (kWh)", "Wind (kWh)",
    "Hydel (kWh)", "Custom Renewable Electricity (kWh)"
}
RENEWABLE_FUEL_TYPES = {"Biofuel (KL)", "Biodiesel (KL)", "Briquettes (Kg)"}
GRID_TYPES = {"Grid Electricity (kWh)"}
FOSSIL_FUEL_TYPES = {
    "HSD (KL)", "HSD Mobile (KL)", "Furnace Oil (KL)", "LDO (KL)",
    "Natural Gas (SCM)", "LPG (Kg)", "LSHS (KL)", "Petrol (KL)"
}


def pct(part, whole):
    return 0.0 if whole == 0 else round((part / whole) * 100, 2)


# ---------------------------------------------------------------------------
# Template Generator
# ---------------------------------------------------------------------------

def generate_template():
    """Create a multi-sheet ESG data-collection template Excel file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    # Styles
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
    period_fill = PatternFill(start_color="155E75", end_color="155E75", fill_type="solid")
    period_font = Font(bold=True, color="FFFFFF", size=11)
    sub_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    sub_font = Font(bold=True, color="065F46", size=10)
    note_font = Font(italic=True, color="6B7280", size=9)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap = Alignment(vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    periods = ["FY 2024-25", "FY 2023-24", "FY 2022-23"]

    def style_header(ws, row, col, value, fill=None, font=None):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = font or hdr_font
        cell.fill = fill or hdr_fill
        cell.alignment = center
        cell.border = thin_border
        return cell

    def style_cell(ws, row, col, value=None, fmt=None):
        cell = ws.cell(row=row, column=col, value=value)
        cell.alignment = center
        cell.border = thin_border
        if fmt:
            cell.number_format = fmt
        return cell

    def add_sample_locations(ws, start_row, loc_col, n=5):
        examples = ["Site 1 - Mumbai", "Site 2 - Delhi", "Site 3 - Bangalore",
                     "Site 4 - Chennai", "Site 5 - Pune"]
        for i, loc in enumerate(examples[:n]):
            cell = ws.cell(row=start_row + i, column=loc_col, value=loc)
            cell.font = Font(italic=True, color="9CA3AF")
            cell.alignment = left_wrap
            cell.border = thin_border

    def auto_width(ws):
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 14), 28)

    # ===== Instructions Sheet =====
    ws0 = wb.active
    ws0.title = "Instructions"
    ws0.sheet_properties.tabColor = "0F766E"

    instructions = [
        ("ESG AutoPilot - Data Collection Template", None, True),
        ("", None, False),
        ("HOW TO USE THIS TEMPLATE", None, True),
        ("1. Each site / location fills in ONE copy of this file.", None, False),
        ("2. Put your site name in the 'Location / Plant' column of each sheet.", None, False),
        ("3. Fill in consumption data under the correct financial year columns.", None, False),
        ("4. Do NOT change the sheet names or column headers.", None, False),
        ("5. Leave cells blank or zero if a fuel type is not used at your site.", None, False),
        ("6. Save the file as:  SiteName_ESG_Data.xlsx", None, False),
        ("", None, False),
        ("UPLOADING", None, True),
        ("All site files can be uploaded together on the ESG AutoPilot dashboard.", None, False),
        ("The system will automatically merge and process all locations.", None, False),
        ("", None, False),
        ("SHEETS IN THIS TEMPLATE", None, True),
        ("Scope 1 Stationary Combustion  -  Boilers, generators, furnaces", None, False),
        ("Scope 1 Mobile Combustion      -  Company-owned vehicles", None, False),
        ("Scope 1 Fugitive Emissions     -  Refrigerant leaks & fire extinguishers", None, False),
        ("Electricity                     -  Grid (NRE) and Renewable (RE) electricity", None, False),
        ("", None, False),
        ("SUPPORTED UNITS", None, True),
        ("Fuel: KL (kilolitres), Kg, SCM (std cubic metres), Litres", None, False),
        ("Electricity: kWh", None, False),
        ("Refrigerants: Kg refilled", None, False),
        ("Fire Extinguishers: Number of units (nos) refilled", None, False),
        ("", None, False),
        ("REPORTING PERIODS", None, True),
        ("Use Indian Financial Year format: FY 2024-25, FY 2023-24, etc.", None, False),
        ("You can add more FY columns by following the same pattern.", None, False),
    ]
    for i, (text, _, is_bold) in enumerate(instructions, 1):
        cell = ws0.cell(row=i, column=1, value=text)
        cell.font = Font(bold=is_bold, size=12 if is_bold else 10,
                         color="0F766E" if is_bold else "374151")
    ws0.column_dimensions["A"].width = 80

    # ===== Scope 1 Stationary Combustion =====
    ws1 = wb.create_sheet("Scope 1 Stationary Combustion")
    ws1.sheet_properties.tabColor = "F59E0B"
    fuels_stat = ["HSD (KL)", "Furnace Oil (KL)", "LDO (KL)",
                  "Natural Gas (SCM)", "LPG (Kg)", "LSHS (KL)"]

    row = 1
    style_header(ws1, row, 1, "Location / Plant")
    col = 2
    for fy in periods:
        style_header(ws1, row, col, fy, period_fill, period_font)
        ws1.merge_cells(start_row=row, start_column=col,
                        end_row=row, end_column=col + len(fuels_stat) - 1)
        col += len(fuels_stat)

    row = 2
    style_header(ws1, row, 1, "")
    col = 2
    for _ in periods:
        for fuel in fuels_stat:
            style_header(ws1, row, col, fuel, sub_fill, sub_font)
            col += 1

    add_sample_locations(ws1, 3, 1, 5)
    for r in range(3, 8):
        for c in range(2, 2 + len(periods) * len(fuels_stat)):
            style_cell(ws1, r, c, fmt="#,##0.00")
    auto_width(ws1)

    # ===== Scope 1 Mobile Combustion =====
    ws2 = wb.create_sheet("Scope 1 Mobile Combustion")
    ws2.sheet_properties.tabColor = "EF4444"
    fuels_mob = ["HSD (KL)", "Petrol (KL)"]

    row = 1
    style_header(ws2, row, 1, "Location / Plant")
    col = 2
    for fy in periods:
        style_header(ws2, row, col, fy, period_fill, period_font)
        ws2.merge_cells(start_row=row, start_column=col,
                        end_row=row, end_column=col + len(fuels_mob) - 1)
        col += len(fuels_mob)

    row = 2
    style_header(ws2, row, 1, "")
    col = 2
    for _ in periods:
        for fuel in fuels_mob:
            style_header(ws2, row, col, fuel, sub_fill, sub_font)
            col += 1

    add_sample_locations(ws2, 3, 1, 5)
    for r in range(3, 8):
        for c in range(2, 2 + len(periods) * len(fuels_mob)):
            style_cell(ws2, r, c, fmt="#,##0.000")
    auto_width(ws2)

    # ===== Scope 1 Fugitive Emissions =====
    ws3 = wb.create_sheet("Scope 1 Fugitive Emissions")
    ws3.sheet_properties.tabColor = "8B5CF6"

    # Refrigerants section
    refrigerants = ["R22 (kg)", "R134a (kg)", "R-407C (kg)", "R-404A (kg)",
                    "R-410A (kg)", "R-407A (kg)", "R123 (kg)", "R32 (kg)", "R152a (kg)"]
    fire_ext = ["1 Kg (nos)", "2 Kg (nos)", "2.5 Kg (nos)", "3 Kg (nos)",
                "3.5 Kg (nos)", "4.5 Kg (nos)", "6 Kg (nos)", "9 Kg (nos)", "22.5 Kg (nos)"]

    row = 1
    ws3.cell(row=row, column=1, value="Refrigerant Gas Refills (kg)").font = Font(
        bold=True, size=12, color="8B5CF6")

    row = 2
    style_header(ws3, row, 1, "Location / Plant")
    col = 2
    for ref in refrigerants:
        style_header(ws3, row, col, ref + " refilled")
        col += 1

    add_sample_locations(ws3, 3, 1, 5)
    for r in range(3, 8):
        for c in range(2, 2 + len(refrigerants)):
            style_cell(ws3, r, c, fmt="#,##0.00")

    fug_note_row = 9
    ws3.cell(row=fug_note_row, column=1,
             value="Note: Enter quantity of refrigerant gas refilled (in kg), "
                   "not the total charge.").font = note_font

    # Fire extinguisher section
    row = fug_note_row + 2
    ws3.cell(row=row, column=1,
             value="CO2 Fire Extinguisher Refills (nos)").font = Font(
        bold=True, size=12, color="8B5CF6")

    row += 1
    style_header(ws3, row, 1, "Location / Plant")
    col = 2
    for fe in fire_ext:
        style_header(ws3, row, col, fe)
        col += 1

    fe_data_start = row + 1
    add_sample_locations(ws3, fe_data_start, 1, 5)
    for r in range(fe_data_start, fe_data_start + 5):
        for c in range(2, 2 + len(fire_ext)):
            style_cell(ws3, r, c, fmt="#,##0")

    fe_note_row = fe_data_start + 6
    ws3.cell(row=fe_note_row, column=1,
             value="Note: Enter number of extinguishers refilled, "
                   "grouped by weight category.").font = note_font
    auto_width(ws3)

    # ===== Electricity =====
    ws4 = wb.create_sheet("Electricity")
    ws4.sheet_properties.tabColor = "3B82F6"

    row = 1
    style_header(ws4, row, 1, "Location / Plant")
    col = 2
    for fy in periods:
        style_header(ws4, row, col, fy, period_fill, period_font)
        ws4.merge_cells(start_row=row, start_column=col,
                        end_row=row, end_column=col + 2)
        col += 3

    row = 2
    style_header(ws4, row, 1, "")
    col = 2
    elec_subs = ["kWh", "NRE", "RE"]
    for _ in periods:
        for sub in elec_subs:
            style_header(ws4, row, col, sub, sub_fill, sub_font)
            col += 1

    add_sample_locations(ws4, 3, 1, 5)
    for r in range(3, 8):
        for c in range(2, 2 + len(periods) * 3):
            style_cell(ws4, r, c, fmt="#,##0")

    elec_note_row = 9
    ws4.cell(row=elec_note_row, column=1,
             value="NRE = Non-Renewable (Grid) Electricity  |  "
                   "RE = Renewable Electricity (Solar, Wind, Hydel, etc.)").font = note_font
    ws4.cell(row=elec_note_row + 1, column=1,
             value="Enter values in kWh. The 'kWh' column is total "
                   "consumption; NRE + RE should equal kWh.").font = note_font
    auto_width(ws4)

    # Save to buffer
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


